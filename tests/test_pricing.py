#!/usr/bin/env python
import unittest
import json
import requests
from openpyxl import load_workbook

COST_CALCULATOR = "tests/test_data/cost_calculator.xlsx"

class TestGet(unittest.TestCase):
    def setUp(self):
        """Server Settings"""
        self.api_url = 'http://localhost:9761/api/v1/'

    def test_all_base_pages(self):
        pages = ['pricing_components', 'pricing_products',
                 'pricing_date_to_version', 'pricing_exchange_rates']

        # Check every url, that it gives a 200 OK response
        error_pages = filter(lambda u: not requests.get(self.api_url + u).ok, pages)

        self.assertTrue(len(error_pages) == 0,
                        msg=('Pages resulted in error: {0} '.format(error_pages)))

    def test_all_detail_pages(self):
        pages = ['pricing_components/6', 'pricing_products/4',
                 'pricing_exchange_rates?date=2018-09-25']

        # Check every url, that it gives a 200 OK response
        error_pages = filter(lambda u: not requests.get(self.api_url + u).ok, pages)

        self.assertTrue(len(error_pages) == 0,
                    msg=('Pages resulted in error: {0} '.format(error_pages)))



    def test_component_prices_against_excel(self):

        def empty_row(row):
            for cell in row:
                if cell.value is not None:
                    return False
            return True

        response = requests.get(self.api_url + 'pricing_components')
        json_data = response.json()

        wb = load_workbook(COST_CALCULATOR, read_only=True, data_only=True)
        ws = wb['Price list']

        for excel_component in ws.iter_rows(min_row=9, max_row=116):
            if empty_row(excel_component):
                continue
            key = excel_component[0].value
            api_component = json_data[str(key)]

            self.assertEqual(api_component['Product name'],
                             excel_component[4].value)

            self.assertAlmostEqual(float(api_component['List price']),
                                   float(excel_component[10].value))

            self.assertAlmostEqual(float(api_component['price_per_unit_in_sek']),
                                   float(excel_component[14].value), places=2)

    def test_product_prices_against_excel(self):

        def empty_row(row):
            for cell in row:
                if cell.value is not None:
                    return False
            return True

        response = requests.get(self.api_url + 'pricing_products')
        json_data = response.json()

        wb = load_workbook(COST_CALCULATOR, read_only=True, data_only=True)
        ws = wb['Products']

        presumed_id = 0
        for excel_product in ws.iter_rows(min_row=4, max_row=59):
            presumed_id += 1
            if empty_row(excel_product):
                continue
            api_product = json_data[str(presumed_id)]

            self.assertEqual(api_product['Name'],
                             excel_product[2].value)

            components = sorted([int(k) for k in api_product['Components'].keys()])
            if len(components) == 1:
                components = components[0]

            excel_components = excel_product[3].value

            # Some cells might be interpreted as floats
            # e.g. "37,78"
            excel_components = str(excel_components)
            excel_components = excel_components.replace('.', ',')
            if excel_components == 'None':
                continue

            if ',' in str(excel_components):
                excel_components = sorted(
                        [int(k) for k in excel_components.split(',')]
                        )
            else:
                excel_components = int(excel_components)

            self.assertEqual(components,
                             excel_components)

            self.assertAlmostEqual(float(api_product['price_internal']),
                                   float(excel_product[6].value), places=2)

            self.assertAlmostEqual(float(api_product['price_external']),
                                   float(excel_product[7].value), places=2)
